# exports/excel_export.py
import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import func
from shared.backend.crud import *
from shared.backend.models.pembayaran import Pembayaran
from shared.backend.models.meteran import Meteran
from dashboard.utils import nama_bulan, rupiah

EXPORT_DIR = os.path.join(os.path.dirname(__file__), "output")
os.makedirs(EXPORT_DIR, exist_ok=True)

BIRU     = "1F4E79"
BIRU_M   = "BDD7EE"
PUTIH    = "FFFFFF"
ABU      = "F2F2F2"

def style_header(cell, bg=None, bold=True, center=True):
    bg = bg or BIRU
    cell.font      = Font(bold=bold, color=PUTIH if bg == BIRU else "000000", size=11)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center")

def border(cell):
    t = Side(style="thin")
    cell.border = Border(left=t, right=t, top=t, bottom=t)

def sheet_ringkasan(wb, db, bulan, tahun):
    ws = wb.create_sheet("Ringkasan Kas", 0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22

    ws.merge_cells("A1:B1")
    ws["A1"] = "RINGKASAN KAS PAMSIMAS TIRTA LESTARI IV"
    style_header(ws["A1"])
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Periode: {nama_bulan(bulan)} {tahun}"
    ws["A2"].font      = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    kas          = sinkron_kas(db, bulan, tahun)
    total_keluar = get_total_pengeluaran_bulan(db, bulan, tahun)
    piutang      = get_piutang_bulan(db, bulan, tahun)
    meteran_bln  = get_meteran_bulanan(db, bulan, tahun)
    jml_lunas    = sum(1 for m, p in meteran_bln
                       if get_sisa_tagihan(db, m.id) == 0)

    rows = [
        ("", ""),
        ("PEMASUKAN", ""),
        ("Target Penarikan",       kas.target_penarikan),
        ("Aktual Terkumpul",       kas.aktual_terkumpul),
        ("Dari Pemasangan Baru",   kas.pemasukan_pemasangan),
        ("Saldo Bulan Lalu",       kas.saldo_bulan_lalu or 0),
        ("", ""),
        ("PENGELUARAN", ""),
        ("Total Pengeluaran",      total_keluar),
        ("", ""),
        ("TOTAL SALDO KAS",        kas.total_saldo),
        ("", ""),
        ("INFO", ""),
        ("Total Pelanggan",        len(meteran_bln)),
        ("Sudah Lunas",            jml_lunas),
        ("Belum Lunas (Piutang)",  len(piutang)),
        ("Dibuat pada", __import__('datetime').datetime.now()
                        .strftime("%d/%m/%Y %H:%M")),
    ]

    for i, (label, nilai) in enumerate(rows, 4):
        ca = ws.cell(row=i, column=1, value=label)
        cb = ws.cell(row=i, column=2, value=nilai)
        if label in ["PEMASUKAN", "PENGELUARAN", "INFO"]:
            style_header(ca, bg=BIRU_M)
            border(ca); border(cb)
        elif label == "TOTAL SALDO KAS":
            style_header(ca); style_header(cb)
            cb.number_format = '#,##0'
            border(ca); border(cb)
        elif nilai != "":
            ca.font = Font(size=11)
            if isinstance(nilai, int):
                cb.number_format = '#,##0'
            border(ca); border(cb)
    return ws

def sheet_tagihan(wb, db, bulan, tahun):
    ws = wb.create_sheet("Rekap Tagihan")
    lebar = [5, 25, 10, 12, 12, 12, 15, 15, 15, 14]
    cols  = ["A","B","C","D","E","F","G","H","I","J"]
    for c, l in zip(cols, lebar):
        ws.column_dimensions[c].width = l

    ws.merge_cells("A1:J1")
    ws["A1"] = f"REKAP TAGIHAN — {nama_bulan(bulan).upper()} {tahun}"
    style_header(ws["A1"])
    ws.row_dimensions[1].height = 28

    headers = ["No","Nama","Area","Desa","Awal(m³)",
               "Akhir(m³)","Kubikasi","Total","Terbayar","Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        style_header(cell); border(cell)

    data  = get_meteran_bulanan(db, bulan, tahun)
    t_tag = t_bayar = 0
    for i, (m, p) in enumerate(data, 1):
        terbayar = get_total_terbayar(db, m.id)
        sisa     = m.total_tagihan - terbayar
        status   = "Lunas" if sisa == 0 else "Belum Lunas"
        row      = 3 + i
        bg       = ABU if i % 2 == 0 else PUTIH
        vals     = [i, p.nama, p.area or "-", p.desa or "-",
                    float(m.angka_awal), float(m.angka_akhir),
                    float(m.kubikasi), m.total_tagihan, terbayar, status]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            border(cell)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center")
            if col in [8, 9]:
                cell.number_format = '#,##0'
            if col == 10:
                cell.font = Font(
                    color="1F7A1F" if status == "Lunas" else "C00000",
                    bold=True
                )
        t_tag   += m.total_tagihan
        t_bayar += terbayar

    tr = 3 + len(data) + 1
    ws.merge_cells(f"A{tr}:G{tr}")
    cell = ws.cell(row=tr, column=1, value="TOTAL")
    style_header(cell, bg=BIRU_M)
    border(cell)
    for col, val in [(8, t_tag), (9, t_bayar)]:
        cell = ws.cell(row=tr, column=col, value=val)
        style_header(cell, bg=BIRU_M)
        cell.number_format = '#,##0'
        border(cell)
    ws.freeze_panes = "A4"
    return ws

def sheet_pembayaran(wb, db, bulan, tahun):
    ws = wb.create_sheet("Pembayaran")
    for col, w in zip(["A","B","C","D","E","F","G","H"],
                      [5, 14, 25, 10, 16, 12, 12, 22]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:H1")
    ws["A1"] = f"PEMBAYARAN — {nama_bulan(bulan).upper()} {tahun}"
    style_header(ws["A1"])
    ws.row_dimensions[1].height = 28

    headers = ["No","Tanggal","Nama","Area","Jumlah","Metode","Status","Catatan"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        style_header(cell); border(cell)

    from shared.backend.models.pembayaran import Pembayaran as PB
    hasil = (
        db.query(PB, Meteran)
        .join(Meteran, PB.meteran_id == Meteran.id)
        .filter(Meteran.bulan == bulan, Meteran.tahun == tahun)
        .order_by(PB.tgl_bayar)
        .all()
    )
    total = 0
    for i, (b, m) in enumerate(hasil, 1):
        row = 3 + i
        bg  = ABU if i % 2 == 0 else PUTIH
        vals = [i, b.tgl_bayar.strftime("%d/%m/%Y"),
                m.pelanggan.nama, m.pelanggan.area or "-",
                b.jumlah_bayar, b.metode, b.status,
                b.catatan or "-"]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            border(cell)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center")
            if col == 5:
                cell.number_format = '#,##0'
        total += b.jumlah_bayar

    tr = 3 + len(hasil) + 1
    ws.merge_cells(f"A{tr}:D{tr}")
    cell = ws.cell(row=tr, column=1, value="TOTAL")
    style_header(cell, bg=BIRU_M); border(cell)
    cell = ws.cell(row=tr, column=5, value=total)
    style_header(cell, bg=BIRU_M)
    cell.number_format = '#,##0'; border(cell)
    ws.freeze_panes = "A4"
    return ws

def sheet_pengeluaran(wb, db, bulan, tahun):
    ws = wb.create_sheet("Pengeluaran")
    for col, w in zip(["A","B","C","D","E"],
                      [5, 14, 30, 14, 25]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E1")
    ws["A1"] = f"PENGELUARAN — {nama_bulan(bulan).upper()} {tahun}"
    style_header(ws["A1"])
    ws.row_dimensions[1].height = 28

    headers = ["No","Tanggal","Keperluan","Jumlah","Direquest Oleh"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        style_header(cell); border(cell)

    data  = get_pengeluaran_bulan(db, bulan, tahun)
    total = 0
    for i, p in enumerate(data, 1):
        row = 3 + i
        bg  = ABU if i % 2 == 0 else PUTIH
        vals = [i, p.tanggal.strftime("%d/%m/%Y"),
                p.keperluan, p.jumlah,
                p.direquest_oleh or "-"]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            border(cell)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center")
            if col == 4:
                cell.number_format = '#,##0'
        total += p.jumlah

    tr = 3 + len(data) + 1
    ws.merge_cells(f"A{tr}:C{tr}")
    cell = ws.cell(row=tr, column=1, value="TOTAL PENGELUARAN")
    style_header(cell, bg=BIRU_M); border(cell)
    cell = ws.cell(row=tr, column=4, value=total)
    style_header(cell, bg=BIRU_M)
    cell.number_format = '#,##0'; border(cell)
    ws.freeze_panes = "A4"
    return ws

def export_laporan_excel(db: Session, bulan: int, tahun: int) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    sheet_ringkasan(wb, db, bulan, tahun)
    sheet_tagihan(wb, db, bulan, tahun)
    sheet_pembayaran(wb, db, bulan, tahun)
    sheet_pengeluaran(wb, db, bulan, tahun)

    filename = f"Laporan_PamSR_{nama_bulan(bulan)}_{tahun}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath

def export_laporan_range(db: Session, 
                          bulan_awal: int, tahun_awal: int,
                          bulan_akhir: int, tahun_akhir: int) -> str:
    """export laporan rentang waktu — semua data dalam 1 sheet"""
    wb = Workbook()
    wb.remove(wb.active)

    # --- SHEET RINGKASAN ---
    ws_ring = wb.create_sheet("Ringkasan", 0)
    ws_ring.column_dimensions["A"].width = 30
    ws_ring.column_dimensions["B"].width = 22

    ws_ring.merge_cells("A1:B1")
    ws_ring["A1"] = "RINGKASAN KAS PAMSIMAS TIRTA LESTARI IV"
    style_header(ws_ring["A1"])
    ws_ring.row_dimensions[1].height = 28

    ws_ring.merge_cells("A2:B2")
    ws_ring["A2"] = (f"Periode: {nama_bulan(bulan_awal)} {tahun_awal} — "
                     f"{nama_bulan(bulan_akhir)} {tahun_akhir}")
    ws_ring["A2"].font      = Font(bold=True, size=11)
    ws_ring["A2"].alignment = Alignment(horizontal="center")

    # hitung semua bulan dalam range
    periode_list = []
    b, t = bulan_awal, tahun_awal
    while (t * 12 + b) <= (tahun_akhir * 12 + bulan_akhir):
        periode_list.append((b, t))
        b += 1
        if b > 12:
            b = 1
            t += 1

    # akumulasi total
    grand_target = grand_terkumpul = grand_masuk = grand_keluar = 0

    ring_rows = [("Bulan", "Target", "Terkumpul", "Pemasangan", "Pengeluaran", "Saldo")]
    for b, t in periode_list:
        kas   = sinkron_kas(db, b, t)
        keluar = get_total_pengeluaran_bulan(db, b, t)
        ring_rows.append((
            f"{nama_bulan(b)} {t}",
            kas.target_penarikan,
            kas.aktual_terkumpul,
            kas.pemasukan_pemasangan,
            keluar,
            kas.total_saldo
        ))
        grand_target   += kas.target_penarikan
        grand_terkumpul += kas.aktual_terkumpul
        grand_masuk    += kas.pemasukan_pemasangan
        grand_keluar   += keluar

    # tulis ringkasan per bulan
    ws_ring.column_dimensions["A"].width = 18
    ws_ring.column_dimensions["B"].width = 16
    ws_ring.column_dimensions["C"].width = 16
    ws_ring.column_dimensions["D"].width = 16
    ws_ring.column_dimensions["E"].width = 16
    ws_ring.column_dimensions["F"].width = 16

    for col, h in enumerate(ring_rows[0], 1):
        cell = ws_ring.cell(row=4, column=col, value=h)
        style_header(cell); border(cell)

    for i, row in enumerate(ring_rows[1:], 1):
        r   = 4 + i
        bg  = ABU if i % 2 == 0 else PUTIH
        for col, val in enumerate(row, 1):
            cell = ws_ring.cell(row=r, column=col, value=val)
            border(cell)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center")
            if col > 1:
                cell.number_format = '#,##0'

    # baris total
    tr = 4 + len(ring_rows)
    ws_ring.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    for col, val in enumerate([
        grand_target, grand_terkumpul,
        grand_masuk, grand_keluar, ""
    ], 2):
        cell = ws_ring.cell(row=tr, column=col, value=val)
        style_header(cell, bg=BIRU_M)
        if isinstance(val, int):
            cell.number_format = '#,##0'
        border(cell)

    ws_ring.freeze_panes = "A5"

    # --- SHEET TAGIHAN (semua bulan digabung) ---
    ws_tag = wb.create_sheet("Rekap Tagihan")
    ws_tag.column_dimensions["A"].width = 5
    ws_tag.column_dimensions["B"].width = 16
    ws_tag.column_dimensions["C"].width = 25
    ws_tag.column_dimensions["D"].width = 10
    ws_tag.column_dimensions["E"].width = 12
    ws_tag.column_dimensions["F"].width = 12
    ws_tag.column_dimensions["G"].width = 15
    ws_tag.column_dimensions["H"].width = 15
    ws_tag.column_dimensions["I"].width = 14

    ws_tag.merge_cells("A1:I1")
    ws_tag["A1"] = (f"REKAP TAGIHAN — "
                    f"{nama_bulan(bulan_awal)} {tahun_awal} s/d "
                    f"{nama_bulan(bulan_akhir)} {tahun_akhir}")
    style_header(ws_tag["A1"])
    ws_tag.row_dimensions[1].height = 28

    headers = ["No","Bulan","Nama","Area","Kubikasi",
               "Total","Terbayar","Sisa","Status"]
    for col, h in enumerate(headers, 1):
        cell = ws_tag.cell(row=3, column=col, value=h)
        style_header(cell); border(cell)

    no = 1
    t_tag = t_bayar = 0
    for b, t in periode_list:
        data = get_meteran_bulanan(db, b, t)
        for m, p in data:
            terbayar = get_total_terbayar(db, m.id)
            sisa     = m.total_tagihan - terbayar
            status   = "Lunas" if sisa == 0 else "Belum Lunas"
            row      = 3 + no
            bg       = ABU if no % 2 == 0 else PUTIH
            vals     = [
                no, f"{nama_bulan(b)} {t}",
                p.nama, p.area or "-",
                float(m.kubikasi),
                m.total_tagihan, terbayar, sisa, status
            ]
            for col, val in enumerate(vals, 1):
                cell = ws_tag.cell(row=row, column=col, value=val)
                border(cell)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center")
                if col in [6, 7, 8]:
                    cell.number_format = '#,##0'
                if col == 9:
                    cell.font = Font(
                        color="1F7A1F" if status=="Lunas" else "C00000",
                        bold=True
                    )
            t_tag   += m.total_tagihan
            t_bayar += terbayar
            no += 1

    tr = 3 + no
    ws_tag.merge_cells(f"A{tr}:E{tr}")
    cell = ws_tag.cell(row=tr, column=1, value="TOTAL")
    style_header(cell, bg=BIRU_M); border(cell)
    for col, val in [(6, t_tag), (7, t_bayar), (8, t_tag-t_bayar)]:
        cell = ws_tag.cell(row=tr, column=col, value=val)
        style_header(cell, bg=BIRU_M)
        cell.number_format = '#,##0'; border(cell)
    ws_tag.freeze_panes = "A4"

    # --- SHEET PEMBAYARAN ---
    ws_bayar = wb.create_sheet("Pembayaran")
    for col, w in zip(["A","B","C","D","E","F","G","H","I"],
                      [5, 16, 14, 25, 10, 16, 10, 12, 22]):
        ws_bayar.column_dimensions[col].width = w

    ws_bayar.merge_cells("A1:I1")
    ws_bayar["A1"] = (f"PEMBAYARAN — "
                      f"{nama_bulan(bulan_awal)} {tahun_awal} s/d "
                      f"{nama_bulan(bulan_akhir)} {tahun_akhir}")
    style_header(ws_bayar["A1"])
    ws_bayar.row_dimensions[1].height = 28

    headers = ["No","Bulan","Tanggal","Nama","Area",
               "Jumlah","Diskon","Metode","Status"]
    for col, h in enumerate(headers, 1):
        cell = ws_bayar.cell(row=3, column=col, value=h)
        style_header(cell); border(cell)

    from shared.backend.models.pembayaran import Pembayaran as PB
    no = 1; total_bayar = total_diskon = 0
    for b, t in periode_list:
        hasil = (
            db.query(PB, Meteran)
            .join(Meteran, PB.meteran_id == Meteran.id)
            .filter(Meteran.bulan == b, Meteran.tahun == t)
            .order_by(PB.tgl_bayar)
            .all()
        )
        for pay, met in hasil:
            row = 3 + no
            bg  = ABU if no % 2 == 0 else PUTIH
            vals = [
                no, f"{nama_bulan(b)} {t}",
                pay.tgl_bayar.strftime("%d/%m/%Y"),
                met.pelanggan.nama, met.pelanggan.area or "-",
                pay.jumlah_bayar, pay.diskon or 0,
                pay.metode, pay.status
            ]
            for col, val in enumerate(vals, 1):
                cell = ws_bayar.cell(row=row, column=col, value=val)
                border(cell)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center")
                if col in [6, 7]:
                    cell.number_format = '#,##0'
            total_bayar  += pay.jumlah_bayar
            total_diskon += pay.diskon or 0
            no += 1

    tr = 3 + no
    ws_bayar.merge_cells(f"A{tr}:E{tr}")
    cell = ws_bayar.cell(row=tr, column=1, value="TOTAL")
    style_header(cell, bg=BIRU_M); border(cell)
    for col, val in [(6, total_bayar), (7, total_diskon)]:
        cell = ws_bayar.cell(row=tr, column=col, value=val)
        style_header(cell, bg=BIRU_M)
        cell.number_format = '#,##0'; border(cell)
    ws_bayar.freeze_panes = "A4"

    # --- SHEET PENGELUARAN ---
    ws_kel = wb.create_sheet("Pengeluaran")
    for col, w in zip(["A","B","C","D","E","F"],
                      [5, 14, 14, 30, 14, 20]):
        ws_kel.column_dimensions[col].width = w

    ws_kel.merge_cells("A1:F1")
    ws_kel["A1"] = (f"PENGELUARAN — "
                    f"{nama_bulan(bulan_awal)} {tahun_awal} s/d "
                    f"{nama_bulan(bulan_akhir)} {tahun_akhir}")
    style_header(ws_kel["A1"])
    ws_kel.row_dimensions[1].height = 28

    headers = ["No","Bulan","Tanggal","Keperluan","Jumlah","Direquest"]
    for col, h in enumerate(headers, 1):
        cell = ws_kel.cell(row=3, column=col, value=h)
        style_header(cell); border(cell)

    no = 1; total_kel = 0
    for b, t in periode_list:
        data = get_pengeluaran_bulan(db, b, t)
        for pen in data:
            row = 3 + no
            bg  = ABU if no % 2 == 0 else PUTIH
            vals = [
                no, f"{nama_bulan(b)} {t}",
                pen.tanggal.strftime("%d/%m/%Y"),
                pen.keperluan, pen.jumlah,
                pen.direquest_oleh or "-"
            ]
            for col, val in enumerate(vals, 1):
                cell = ws_kel.cell(row=row, column=col, value=val)
                border(cell)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="center")
                if col == 5:
                    cell.number_format = '#,##0'
            total_kel += pen.jumlah
            no += 1

    tr = 3 + no
    ws_kel.merge_cells(f"A{tr}:D{tr}")
    cell = ws_kel.cell(row=tr, column=1, value="TOTAL PENGELUARAN")
    style_header(cell, bg=BIRU_M); border(cell)
    cell = ws_kel.cell(row=tr, column=5, value=total_kel)
    style_header(cell, bg=BIRU_M)
    cell.number_format = '#,##0'; border(cell)
    ws_kel.freeze_panes = "A4"

    # simpan file
    filename = (f"Laporan_PamSR_"
                f"{nama_bulan(bulan_awal)}{tahun_awal}_"
                f"sd_{nama_bulan(bulan_akhir)}{tahun_akhir}.xlsx")
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath
